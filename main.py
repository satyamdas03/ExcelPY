from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')

# ws = wb.active
# ws['A2'].value = "Test"
# print(wb.sheetnames)
# wb.save('Grades.xlsx')
wb.create_sheet("Test")
print(wb.sheetnames)
# rest to start shortly
# creating new workbook
#adding/appending rows
# accessing multiple cells
# merging cells
# inserting and deleting rows
#copying and moving cells
#practical example and formula for moving cells and styling them